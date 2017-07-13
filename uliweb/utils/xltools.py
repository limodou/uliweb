#coding=utf-8
#xlsx模板处理

import os
import re
import shutil
from openpyxl import load_workbook, Workbook, __version__
from pprint import pprint
from copy import deepcopy
import datetime
from uliweb.utils.common import safe_unicode

ver_4 = __version__ > '2.3.5'

re_filter = re.compile(r'(\w+)(?:\((.*?)\))?')

def _date(cell, data, f, _var=None, _sh=None):
    f = f or 'yyyy-mm-dd'
    cell.value = data
    cell.number_format = f
    return cell.value

def _link(cell, data, text, _var=None, _sh=None):
    t = safe_unicode(text).format(**_var)
    cell.value = data
    cell.hyperlink = t
    return cell.value

def _validate_list(cell, data, _list, blank=True, _var=None, _sh=None):
    from openpyxl.worksheet.datavalidation import DataValidation
    cell.value = data
    dv = DataValidation(type="list", formula1='"{}"'.format(_list), allow_blank=blank)
    _sh.add_data_validation(dv)
    dv.add(cell)
    return cell.value

def hyperlink(title, link, domain=None):
    import urlparse

    r = urlparse.urlparse(link)
    if not r.scheme and domain:
        link = domain + link
    title = title.replace('"', '')
    return title, link

env = {'date':_date, 'datetime':datetime, 'link':_link,
       'validate_list':_validate_list}

def get_line(sheet, line):
    if ver_4:
        return sheet[line]
    else:
        return sheet.rows[line-1]

class Converter(object):
    """
    Parse expr|filter|filter(args)
    """
    def __init__(self, text):
        self.text = text
        self.expr = ''
        self.filter = []
        self.parse()

    def parse(self):
        for i, c in enumerate(self.text.split('|')):
            c = c.strip()
            if i == 0:
                self.expr = c
            else:
                if '(' in c:
                    b = re_filter.match(c)
                    if b:
                        f = u'{0}(cell, value, {1}, _var=_var, _sh=_sh)'.format(b.group(1), b.group(2))
                    else:
                        raise ValueError("Filter {} format is not right.".format(c))
                else:
                    f = u'{}(cell, value, _var=_var, _sh=_sh)'.format(c)
                self.filter.append(f)

    def __call__(self, cell, data, env, sheet=None):
        from openpyxl.styles import Font

        d = data.copy()
        d['cell'] = cell

        if self.expr:
            value = eval(self.expr.replace('.', '_'), env, data)
        else:
            value = self.expr

        d['value'] = value
        d['_var'] = d
        d['_sh'] = sheet
        cell.value = value

        #处理链接
        r = None
        if isinstance(value, (str, unicode)):
            r = re_link.search(value)
            if r:
                value, link = hyperlink(r.group(2), r.group(1))
        if r:
            font = Font(underline='single', color='FF0000FF')
            cell.font = font
            cell.hyperlink = link
            cell.value = value


        for f in self.filter:
            d['value'] = eval(f, d, env)
        return d['value']

re_link = re.compile('^<a.*?href=\"([^\"]+)\".*?>(.*?)</a>', re.DOTALL)



def get_range_string(start_column, start_row, end_column, end_row):
    from openpyxl.utils import get_column_letter

    return (get_column_letter(start_column)+str(start_row) + ':' +
            get_column_letter(end_column)+str(end_row))

class SimpleWriter(object):
    def __init__(self, filename=None, sheet_name=None, header=None, data=None,
                 write_header=True, begin_col=1, begin_row=1, merge_fields=None,
                 border=True, border_color=None, header_color=None, domain=None,
                 encoding=None,
                 default_min_width=10, default_max_width=50, auto_width=True):
        self.header = header or []
        self.data = data
        self.filename = filename
        self.sheet_name = sheet_name
        self.write_header = write_header
        self.merge_fields = merge_fields or []
        self.begin_col = begin_col
        self.begin_row = begin_row
        self._row = begin_row
        self._col = begin_col
        self.wb = None
        self.sheet = None
        self.border = border
        self.border_color = border_color or 'FF000000'
        self.header_color = header_color or 'FF33AAFF'
        self.default_min_width = default_min_width
        self.default_max_width = default_max_width
        self.auto_width = auto_width
        self.widths = {}
        self.fields_list = []
        self.fields = {}
        self.field_names = []
        if domain:
            self.domain = domain.rstrip('/')
        else:
            self.domain = domain

    def _write_header(self):
        from uliweb.utils.common import safe_unicode

        fields = []
        max_rowspan = 0
        for i, f in enumerate(self.fields_list):
            _f = list(f['title'].split('/'))
            max_rowspan = max(max_rowspan, len(_f))
            fields.append((_f, i))

        def get_field(fields, i, m_rowspan):
            f_list, col = fields[i]
            field = {'title':f_list[0], 'col':col, 'colspan':1, 'rowspan':1}
            if len(f_list) == 1:
                field['rowspan'] = m_rowspan
            return field

        def remove_field(fields, i):
            del fields[i][0][0]

        def clear_fields(fields):
            n = len(fields)
            for i in range(len(fields)-1, -1, -1):
                if len(fields[i][0]) == 0:
                    n = min(n, fields[i][1])
                    del fields[i]
            if len(fields):
                return fields[0][1]
            return 0

        n = len(fields)
        y = 0
        while n>0:
            i = 0
            while i<n:
                field = get_field(fields, i, max_rowspan-y)
                remove_field(fields, i)
                j = i + 1
                while j<n:
                    field_n = get_field(fields, j, max_rowspan-y)
                    if safe_unicode(field['title']) == safe_unicode(field_n['title']) and field['rowspan'] == field_n['rowspan']:
                        #combine
                        remove_field(fields, j)
                        field['colspan'] += 1
                        j += 1
                    else:
                        break

                _f = self.fields[field['col']]
                y1 = self._row + y
                y2 = y1 + field['rowspan'] - 1
                x1 = self._col + field['col']
                x2 = x1 + field['colspan'] - 1
                cell = self.sheet.cell(column=x1, row=y1, value=safe_unicode(field['title']))
                self.sheet.merge_cells(start_row=y1, end_row=y2,
                                       start_column=x1, end_column=x2)
                self._format_header_cell(cell, start_row=y1, end_row=y2,
                                       start_column=x1, end_column=x2)


                i = j
            clear_fields(fields)
            n = len(fields)
            y += 1

        self._row += y

    def _process_header(self):
        if not self.header:
            return

        for i, field in enumerate(self.header):
            if isinstance(field, dict):
                f = field.copy()
                if 'verbose_name' in field:
                    f['title'] = field['verbose_name']
            else:
                raise ValueError("Dict type should be used, but {!r} found".format(field))
            self.fields[i] = f
            self.fields_list.append(f)
            self.field_names.append(f['name'])

    def _set_width(self, col, value):
        from uliweb.utils.common import safe_unicode

        length = len(safe_unicode(value))
        if self.auto_width:
            self.widths[col] = min(max(self.default_min_width, length), self.default_max_width)

    def _set_widths(self):
        if self.auto_width:
            from openpyxl.utils import get_column_letter

            col = self._col
            #设置列宽度
            for i in range(len(self.header)):
                self.sheet.column_dimensions[get_column_letter(col+i)].width = self.widths[col+i]

    def _write_body(self):
        x = self._col
        y = self._row
        j = 0
        #存储将要合并单元格信息
        #{name:{row: col: value}}
        last_merge_data = {}
        for j, row in enumerate(self.data):
            for i, col in enumerate(self.header):
                name = col['name']
                if isinstance(row, (tuple, list)):
                    value = row[i]
                elif isinstance(row, dict):
                    value = row[name]
                else:
                    raise ValueError("Row data should be tuple, list or dict, but {!r} found".format(row))

                #处理链接
                r = None
                if isinstance(value, (str, unicode)):
                    r = re_link.search(value)
                    if r:
                        value, link = hyperlink(r.group(2), r.group(1), self.domain)
                cell = self.sheet.cell(column=x+i, row=y+j, value=value)
                if r:
                    cell.hyperlink = link

                #处理列宽度
                self._set_width(x+i, value)

                #处理样式
                self._format_cell(cell, i)

                #处理单元格合并
                if self.merge_fields and name in self.merge_fields:
                    if name in last_merge_data:
                        v = last_merge_data[name]
                        if v['value'] != value: #相同则忽略,不相同则合并
                            #合并当前单元格之后的（包括当前单元格）
                            last_row = y + j - 1
                            ix = self.merge_fields.index(name)
                            self._merge_cells(last_row, ix, last_merge_data)
                            last_merge_data[name] = {'row':y+j, 'col':x+i, 'value':value}
                        else:
                            continue
                    else:
                        last_merge_data[name] = {'row':y+j, 'col':x+i, 'value':value}


        last_row = y+j
        for k, v in last_merge_data.items():
            if last_row > v['row']:
                self.sheet.merge_cells(start_row=v['row'],
                                                    start_column=v['col'],
                                                    end_row=last_row,
                                                    end_column=v['col'])

        self._set_widths()

    def _merge_cells(self, row, index, last_merge_data):
        # flag = False
        for i in range(index, len(self.merge_fields)):
            name = self.merge_fields[i]
            v = last_merge_data[name]
            if row > v['row']:
                self.sheet.merge_cells(start_row=v['row'],
                                    start_column=v['col'],
                                    end_row=row,
                                    end_column=v['col'])
            del last_merge_data[name]

    def _style_range(self, cell, cell_range, border=None, fill=None, font=None, alignment=None):
        """
        Apply styles to a range of cells as if they were a single cell.

        :param ws:  Excel worksheet instance
        :param range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border
        :param fill: An openpyxl PatternFill or GradientFill
        :param font: An openpyxl Font object
        """

        from openpyxl.styles import Border, Side

        top = left = right = bottom = Side(border_style='thin', color=self.border_color)

        def border_add(border, top=None, right=None, left=None, bottom=None):
            top = top or border.top
            left = left or border.left
            right = right or border.right
            bottom = bottom or border.bottom
            return Border(top=top, left=left, right=right, bottom=bottom)

        cell.alignment = alignment
        cell.fill = fill

        rows = list(self.sheet[cell_range])

        for cell in rows[0]:
            cell.border = border_add(cell.border, top=top)
        for cell in rows[-1]:
            cell.border = border_add(cell.border, bottom=bottom)

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = border_add(l.border, left=left)
            r.border = border_add(r.border, right=right)
            # if fill:
            #     for c in row:
            #         c.fill = fill

    def _format_header_cell(self, cell, start_column, start_row, end_column, end_row):
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

        align = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(start_color=self.header_color,
               end_color=self.header_color,
               fill_type='solid')
        border = Border(left=Side(border_style='thin',
                       color=self.border_color),
             right=Side(border_style='thin',
                        color=self.border_color),
             top=Side(border_style='thin',
                      color=self.border_color),
             bottom=Side(border_style='thin',
                         color=self.border_color))

        self._style_range(cell,
                    get_range_string(start_column, start_row, end_column, end_row),
                    border=border, fill=fill, alignment=align)


    def _format_cell(self, cell, index):
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

        head = self.header[index]
        if 'number_format' in head:
            cell.number_format = head['number_format']
        if 'align' in head:
            align = Alignment(horizontal=head['align'],
                     vertical=head.get('valign', 'center'))
            cell.alignment = align
        if self.border:
            border = Border(left=Side(border_style='thin',
                           color=self.border_color),
                 right=Side(border_style='thin',
                            color=self.border_color),
                 top=Side(border_style='thin',
                          color=self.border_color),
                 bottom=Side(border_style='thin',
                             color=self.border_color))
            cell.border = border
        if cell.hyperlink:
            font = Font(underline='single', color='FF0000FF')
            cell.font = font


    def _write(self):
        if self.filename and os.path.exists(self.filename):
            wb = load_workbook(self.filename, keep_vba=True)
        else:
            wb = Workbook()
        self.wb = wb
        if self.sheet_name:
            if self.sheet_name in wb.sheetnames:
                sheet = wb[self.sheet_name]
            else:
                sheet = wb.create_sheet(title=self.sheet_name)
        else:
            sheet = wb.active

        self.sheet = sheet
        if self.write_header:
            self._process_header()
            self._write_header()
        self._write_body()

    def save(self, filename=None):
        self._write()
        self.wb.save(filename or self.filename)


class WriteTemplate(object):
    def __init__(self, sheet):
        self.sheet = sheet
        self.size = (sheet.max_row, sheet.max_column)
        self.row = 0
        self.begin = 1 #记录第一个for的位置,用在简单模板读取中
        self.template = self.get_template()

    def get_style(self, cell):
        return {'font':cell.font, 'border':cell.border, 'fill':cell.fill,
                 'alignment':cell.alignment, 'protection':cell.protection,
                 'number_format':cell.number_format,
                 }

    def set_style(self, cell, style):
        for k, v in style.items():
            setattr(cell, k, v)

    def get_template(self):
        """
        读取一个Excel模板，将此Excel的所有行读出来，并且识别特殊的标记进行记录

        :return: 返回读取后的模板，结果类似：

        [
            {'cols': #各列,与subs不会同时生效
             'subs':[ #子模板
                    {'cols':#各列,
                     'subs': #子模板
                     'field': #对应数据中字段名称
                    },
                    ...
                ]
             'field': #对应数据中字段名称
            },
            ...
        ]

        子模板的判断根据第一列是否为 {{for field}} 来判断，结束使用 {{end}}
        """
        rows = []
        stack = []
        stack.append(rows)
        #top用来记录当前栈
        top = rows
        for i in range(1, self.sheet.max_row+1):
            cell = self.sheet.cell(row=i, column=1)
            #是否子模板开始
            if (isinstance(cell.value, (str, unicode)) and
                    cell.value.startswith('{{for ') and
                    cell.value.endswith('}}')):
                row = {'field':cell.value[6:-2].strip(), 'cols':[], 'subs':[]}
                top.append(row)
                top = row['subs']
                stack.append(top)
                if self.begin == 1:
                    self.begin = i
            #是否子模板结束
            elif (isinstance(cell.value, (str, unicode)) and
                    cell.value == '{{end}}'):
                stack.pop()
                top = stack[-1]
            else:
                row = {'cols':[], 'subs':[]}
                cols = row['cols']
                for j in range(1, self.sheet.max_column+1):
                    cell = self.sheet.cell(row=i, column=j)
                    v = self.process_cell(i, j, cell)
                    if v:
                        cols.append(v)
                if row['cols'] or row['subs']:
                    top.append(row)


        # pprint(rows)
        return rows

    def parse_cell(self, cell):
        """
        Process cell field, the field format just like {{field}}
        :param cell:
        :return: value, field
        """
        field = ''
        if (isinstance(cell.value, (str, unicode)) and
            cell.value.startswith('{{') and
            cell.value.endswith('}}')):
            field = cell.value[2:-2].strip()
            value = ''
        else:
            value = cell.value

        return value, field

    def process_cell(self, row, column, cell):
        value, field = self.parse_cell(cell)
        if field:
            converter = Converter(field)
        else:
            converter = None

        return {'value':value, 'converter':converter, 'col':column,
                'style':self.get_style(cell)}

    def write(self, sheet, data):
        for v in data:
            self.row += 1
            v['row'] = self.row
            self.write_single(sheet, v)

        i = self.row
        while i < self.size[0] + 1:
            for j in range(self.size[1]):
                cell = sheet.cell(row=i, column=j+1)
                cell.value = None
            i += 1
        # self.sheet._garbage_collect()


    def write_line(self, sheet, row, data):
        """

        :param sheet:
        :param row: template row
        :param data:
        :return:
        """
        for i, d in enumerate(row['cols']):
            c = sheet.cell(row=self.row, column=i+1)
            self.write_cell(sheet, c, d, data)
        self.row += 1

    def write_cell(self, sheet, cell, column, data):
        self.set_style(cell, column['style'])
        if column['converter']:
            value = column['converter'](cell, data, env, sheet)
        else:
            value = column['value']
            cell.value = value

    def write_single(self, sheet, data):
        """

        :param sheet:
        :param data: 报文对象， dict
        :param template:
        :return:
        """
        # if not get_data:
        #     def get_data(x):
        #         # print '=====', x
        #         for i in x:
        #             yield i
        #             if 'children' in i and i['children']:
        #                 for j in get_data(i['children']):
        #                     yield j
        #
        def _subs(sheet, subs, data):
            for d in data:
                if d:
                    for line in subs:
                        self.write_line(sheet, line, d)

        for line in self.template:
            if line['subs']:
                loop_name = line['field']
                d = data.get(loop_name, [{}])
                _subs(sheet, line['subs'], d)
            else:
                self.write_line(sheet, line, data)

class ReadLines(object):
    def __init__(self, sheet, start=1):
        self.line = start
        self.size = sheet.max_row

    def next(self):
        if self.line < self.size:
            self.line += 1
        else:
            self.line = None
        return self.line

class ReadTemplate(WriteTemplate):
    def process_cell(self, row, column, cell):
        """
        对于读模板，只记录格式为 {{xxx}} 的字段
        :param cell:
        :return: 如果不是第一行，则每列返回{'col':列值, 'field'}，否则只返回 {'value':...}
        """
        value, field = self.parse_cell(cell)

        if value or field:
            return {'col':column, 'field':field, 'value':value}
        else:
            return {}

    def match(self, row, template_row=None):
        """
        匹配一个模板时，只比较起始行，未来考虑支持比较关键字段即可，现在是起始行的所有字段全匹配
        :param row:
        :return:
        """
        if not template_row:
            template_cols = self.template[0]['cols']
        else:
            template_cols = template_row['cols']

        #check if length of template_cols great than row, then match failed
        if len(template_cols)>len(row):
            return False

        for c in template_cols:
            #如果坐标大于模板最大长度，则表示不匹配，则退出
            text = c['value']
            if not text or (text and not (text.startswith('{{') and text.endswith('}}'))
                and row[c['col']-1].value == text):
                pass
            else:
                return False
        # print 'Matched'
        return True


    def extract_data(self, sheet, line):

        def _get_line(line, row):
            d = {}
            for c in row['cols']:
                if c['field']:
                    cell = sheet.cell(row=line, column=c['col'])
                    if isinstance(cell.value, (str, unicode)):
                        v = cell.value.strip()
                    else:
                        v = cell.value
                    d[c['field']] = v
            return d

        def _subs(line, subs, nextrow):
            data = []
            while line:
                if not nextrow or (nextrow and not self.match(get_line(sheet, line), nextrow)):
                    d = {}
                    for r in subs:
                        _d = _get_line(line, r)
                        d.update(_d)
                        line = line_iter.next()
                    #skip blank line, all columns value should be None
                    if d and filter(None, d.values()):
                        data.append(d)
                else:
                    break
            return data

        line_iter = ReadLines(sheet, line)
        result = []
        data = {}

        while line_iter.line is not None:
            n = 0
            data = {}
            while line_iter.line and n < len(self.template):
                row = self.template[n]
                if row['subs']:
                    if n == len(self.template) - 1:
                        nextrow = self.template[0]
                    else:
                        nextrow = self.template[n+1]
                    data[row['field']] = _subs(line_iter.line, row['subs'], nextrow)
                else:
                    d = _get_line(line_iter.line, row)
                    if d:
                        data.update(d)
                    line_iter.next()

                n += 1

            if data and filter(None, data.values()):
                result.append(data)

        return result

    def get_data(self, sheet, find=False, begin=1):
        """
        Extract data from sheet
        :param find: 查找模式.缺省为False.为True时,将进行递归查找
        :param begin: 开始行号
        """
        line = begin
        for row in sheet.rows:
            if self.match(row):
                d = self.extract_data(sheet, line)
                return d
            else:
                if find:
                    line += 1
                else:
                    break
        return

    def read_data(self, sheet, begin=None):
        """
        用于简单模板匹配,只处理一行的模板, begin为None时自动从for行开始
        :param sheet: 应该使用read_only模式打开
        :param begin:
        :return:
        """
        line = self.begin
        rows = sheet.rows
        for i in range(line-1):
            rows.next()
        template_line = self.template[self.begin-1]
        if not template_line['subs']:
            raise ValueError("Template definition is not right")

        for row in rows:
            d = {}
            for c in template_line['subs'][0]['cols']:
                if c['field']:
                    cell = row[c['col']-1]
                    if isinstance(cell.value, (str, unicode)):
                        v = cell.value.strip()
                    else:
                        v = cell.value
                    d[c['field']] = v
            yield d

class Writer(object):
    def __init__(self, template_file, sheet_name, output_file, data, create=True):
        self.template_file = template_file
        self.output_file = output_file
        self.sheet_name = sheet_name
        self.data = data
        self.create = create
        self.row = 0

        self.init()

    def get_sheet(self, wb):
        if not self.sheet_name:
            sheet = wb.active
        else:
            sheet = wb[self.sheet_name]
        return sheet

    def init(self):

        if self.create or not os.path.exists(self.output_file):
            shutil.copy(self.template_file, self.output_file)
        wb1 = load_workbook(self.template_file)
        template = self.get_template(wb1)

        wb2 = load_workbook(self.output_file)
        template.write(self.get_sheet(wb2), self.data)

        wb2.save(self.output_file)

    def get_template(self, workbook):
        t = WriteTemplate(self.get_sheet(workbook))
        return t



class Reader(object):
    def __init__(self, template_file, sheet_name, input_file,
                 use_merge=False, merge_keys=None, merge_left_join=True,
                 merge_verbose=False, find=False, begin=1, callback=None):
        """

        :param template_file:
        :param sheet_name:
        :param input_file: 输入文件可以是一个list或tuple数组,如果是tuple数组,格式为:
                [('filename', '*'), ('filename', 'sheetname1', 'sheetname2'), 'filename']
                第一个为文件名,后面的为sheet页名称,'*'表示通配符,如果和sheet_name相同,则可以仅为字符串

                如果input_file中不存在指定的Sheet,则自动取所有sheets
        :param use_merge: If use Merge to combine multiple data
        :param merge_keys: keys parameter used in Merge
        :param merge_left_join: left_join parameter used in Merge
        :param merge_verbose: verbose parameter used in Merge
        :param find: 是否查找模板,缺省为False
        :param begin: 是否从begin行开始,缺省为1
        :param callback: callback function used after combine data, the callback function
                should be:

                ```
                def func(row_data):
                ```
        :return:
        """
        self.template_file = template_file
        self.sheet_name = sheet_name
        self.matched = False #if matched template
        if isinstance(input_file, (str, unicode)):
            self.input_file = [input_file]
        elif isinstance(input_file, (tuple, list)):
            self.input_file = input_file
        else:
            raise ValueError("input_file parameter should be tuple or list, but {!r} found".format(type(input_file)))
        self.result = None
        self.callback = callback
        if use_merge:
            self.merge = Merge(keys=merge_keys, left_join=merge_left_join, verbose=merge_verbose)
        else:
            self.merge = False

        template = self.get_template()

        self.result = self.read(template, find=find, begin=begin)

        if self.callback:
            self.callback(self.result)

    def get_template(self):
        wb = load_workbook(self.template_file)

        sheet = wb[self.sheet_name]
        t = ReadTemplate(sheet)
        return t

    def get_sheet(self):
        for f in self.input_file:
            if isinstance(f, (str, unicode)):
                w = load_workbook(f)
                if self.sheet_name not in w.sheetnames:
                    sheet_names = w.sheetnames
                else:
                    sheet_names = [self.sheet_name]
            elif isinstance(f, (tuple, list)):
                if len(f) <= 1:
                    raise ValueError("Filename {} should be followed sheetnames, "
                        "but only filename found".format(f[0]))
                w = load_workbook(f[0])
                if f[1] == '*':
                    sheet_names = w.sheetnames
                else:
                    sheet_names = f[1:]
            else:
                raise ValueError("input_file argument should be string list "
                                 "or tuple list, but {!r} found".format(self.input_file))

            #返回sheet对象
            for name in sheet_names:
                yield w[name]


    def read(self, template, find=False, begin=1):
        """
        :param find: 是否使用find模式.True为递归查找.缺省为False
        :param begin: 开始行号. 缺省为 1
        """
        result = []
        self.matched = False
        for sheet in self.get_sheet():
            r = template.get_data(sheet, find=find, begin=begin)
            if r is not None: #not matched
                self.matched = True
                if self.merge:
                    self.merge.add(r)
                    result = self.merge.result
                else:
                    result.extend(r)
        return result

class SimpleReader(object):
    def __init__(self, template_file, input_file, sheet_name=None,
                 data_sheet_name=None,
                 begin=None):
        """
        只用来处理简单读取,即数据为单行的模式
        :param template_file:
        :param input_file: 输入文件名
        :param sheet_name: 当只有一个sheet时,不管名字对不对都进行处理
        :param find: 是否查找模板,缺省为False
        :param begin: 是否从begin行开始,缺省为None,表示自动根据模板进行判断
        :return:
        """
        self.template_file = template_file
        self.sheet_name = sheet_name
        self.input_file = input_file
        self.begin = begin
        self.data_sheet_name = data_sheet_name or sheet_name
        self.template = self.get_template()

    def get_template(self):
        wb = load_workbook(self.template_file)

        if not self.sheet_name:
            sheet = wb.active
        else:
            sheet = wb[self.sheet_name]
        t = ReadTemplate(sheet)
        return t

    def get_sheet(self):
        w = load_workbook(self.input_file, read_only=True)
        if self.data_sheet_name:
            if self.data_sheet_name not in w.sheetnames:
                raise ValueError("Sheet name [{}] is not found in file {}".format(self.data_sheet_name, self.input_file))
            if self.data_sheet_name == '*':
                for x in w.sheetnames:
                    yield w.sheetnames[x]
            else:
                yield w[self.data_sheet_name]
        else:
            yield w.active

    def read(self):
        """
        :param find: 是否使用find模式.True为递归查找.缺省为False
        :param begin: 开始行号. 缺省为 None, 表示使用模板计算的位置
        """
        for sheet in self.get_sheet():
            for row in self.template.read_data(sheet, begin=self.begin):
                yield row

class Merge(object):
    def __init__(self, keys=None, left_join=True, verbose=False):
        """
        :param alist: 将要合并的数组,每个元素为一个字典
        :param keys: 指明字典的key,如果是多层,则为路径,值为list,如 ['key', 'key1/key2']
        :return:
        """
        self.result = []
        self.keys = keys or []
        self.left_join = left_join
        self.verbose = verbose

    def add(self, *alist):
        for a in alist:
            if self.result:
                self.result = self.merge_list(self.result, deepcopy(a))
            else:
                self.result = deepcopy(a)

    def get_compare_key(self, d, path):
        v = []
        for k in d.keys():
            if path:
                _k = path + '/' + k
            else:
                _k = k
            if _k in self.keys:
                v.append(d[k])
        if v:
            return tuple([_k] + v)
        else:
            return v

    def merge_list(self, list1, list2, path=''):
        objs = {}
        result = []
        for i in list1:
            _key = self.get_compare_key(i, path)
            if _key:
                objs[_key] = i
                result.append(i)
            else:
                if i not in result:
                    result.append(i)
        for i in list2:
            _key = self.get_compare_key(i, path)
            if _key:
                if _key not in objs:
                    if not self.left_join:
                        d = deepcopy(i)
                        result.append(d)
                        objs[_key] = d
                    #如果是left_join=True,以左则数组为准
                    else:
                        if self.verbose:
                            print '---- Path=', path, 'Skip=', i
                else:
                    self.merge_dict(objs[_key], i)
            else:
                #如果是第二层以下,则简单合并,否则进行字典合并
                if path:
                    if not i in result:
                        result.append(i)
                else:
                    self.merge_dict(result[0], i)
                # if i not in result:
                #     result.append(i)

        return result

    def merge_dict(self, d1, d2):
        for k, v in d1.items():
            if isinstance(v, list):
                d1[k] = self.merge_list(v, d2.get(k, []), path=k)
            elif isinstance(v, dict):
                d1[k] = self.merge_dict(v, d2.get(k, {}))
            else:
                d1[k] = d2.get(k, d1[k])

if __name__ == '__main__':
    header = [{'width':10, 'title':'名字', 'name':'name', 'align':'center'},
        {'width':10, 'title':'信息/人员/年龄', 'align':'left', 'name':'age'},
        {'width':10, 'title':'信息/人员/aaaa', 'name':'aaaa', 'align':'right'},
        {'width':20, 'title':'信息/日期', 'number_format':'yyyy-mm-dd', 'name':'date'},
        {'title':'链接', 'name':'link'}]
    data = [
        ('中文', 12, 12.3, datetime.datetime(2011, 7, 11), '<a href="http://google.com">google.com</a>'),
        {'name':'中文', 'age':12, 'aaaa':15, 'date':datetime.datetime(2015, 7, 11),
         'link':'<a href="http://google.com">google.com</a>'},
        {'name':'中文', 'age':13, 'aaaa':15, 'date':datetime.datetime(2015, 7, 11),
         'link':'<a href="http://google.com">google.com</a>'},
        {'name':'中文', 'age':13, 'aaaa':15, 'date':datetime.datetime(2015, 7, 11),
         'link':'<a href="http://google.com">google.com</a>'},
        {'name':'中文1', 'age':13, 'aaaa':15, 'date':datetime.datetime(2015, 7, 11),
         'link':'<a href="http://google.com">google.com</a>'},
        {'name':'中文1', 'age':14, 'aaaa':15, 'date':datetime.datetime(2015, 7, 11),
         'link':'<a href="http://google.com">google.com</a>'},
    ]
    w = SimpleWriter(header=header, data=data, merge_fields=['name', 'age'])
    w.save('test.xlsx')